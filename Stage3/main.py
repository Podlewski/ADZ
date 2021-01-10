import pandas as pd
from openpyxl import Workbook
from tweet_similarity import create_tweet_similarity_sheet
from fuzzy_sets import create_fuzzy_sheet, create_fuzzy_interval_sheet, create_r_coefficient


def freeze(sheet):
    sheet.freeze_panes = sheet["B2"]


workbook = Workbook()
workbook.remove(workbook.active)
tweets_similarity = workbook.create_sheet("Tweets similarity")
fuzzy_membership = workbook.create_sheet("Fuzzy membership")
fuzzy_interval_membership = workbook.create_sheet("Fuzzy interval membership")
r_coef = workbook.create_sheet("Receival to popularity R coef")

df = pd.read_csv("Climate_twitter.csv", sep=',', usecols=['twitter_name', 'followers', 'likes', 'text', 'polarity'])

print("## Creating tweet similarity sheet")
create_tweet_similarity_sheet(tweets_similarity, df.text)
print("## Creating fuzzy sheet")
create_fuzzy_sheet(fuzzy_membership, df)
print("## Creating fuzzy interval sheet")
create_fuzzy_interval_sheet(fuzzy_interval_membership, df)
print("## Creating r coefficient sheet")
create_r_coefficient(r_coef, df)

freeze(tweets_similarity)
freeze(fuzzy_membership)
freeze(fuzzy_interval_membership)
freeze(r_coef)

workbook.save("adz3.xlsx")